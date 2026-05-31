"""Tests for workbook generation (U7).

These tests use a temporary results.db populated with minimal fixture data so
they run offline without DATABASE_URL.
"""

import sqlite3
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from workbook.generator import generate_workbook
from workbook.styles import TAB_NAMES

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

_EXPECTED_SHEETS = list(TAB_NAMES)


def _make_results_db(path: Path) -> None:
    """Populate a minimal results.db with one row per table."""
    conn = sqlite3.connect(path)
    conn.execute("""
        CREATE TABLE results_net_revenue (
            retailer TEXT, gross_revenue REAL, trade_spend REAL,
            net_revenue REAL, net_to_gross_ratio REAL
        )
    """)
    conn.execute(
        "INSERT INTO results_net_revenue VALUES (?,?,?,?,?)",
        ("Walmart", 1000000.0, 120000.0, 880000.0, 0.88),
    )
    conn.execute("""
        CREATE TABLE results_leakage_summary (
            leakage_type TEXT, display_name TEXT, dollar_total REAL,
            instance_count INTEGER, classification TEXT
        )
    """)
    conn.execute(
        "INSERT INTO results_leakage_summary VALUES (?,?,?,?,?)",
        ("double_funded", "Double-Funded", 15264.0, 173, "Recoverable"),
    )
    conn.execute("""
        CREATE TABLE results_leakage_instances (
            leakage_type TEXT, deduction_id TEXT, retailer_id TEXT,
            promo_id TEXT, period TEXT, agreed_amount REAL,
            actual_amount REAL, variance REAL, classification TEXT
        )
    """)
    conn.execute(
        "INSERT INTO results_leakage_instances VALUES (?,?,?,?,?,?,?,?,?)",
        ("double_funded", "DED-001", "walmart", "PROMO-01", "2024-01", 500.0, 600.0, -100.0, "Recoverable"),
    )
    conn.execute("""
        CREATE TABLE results_trade_efficiency (
            retailer TEXT, trade_spend_pct REAL, trade_spend REAL,
            gross_revenue REAL, total_promo_cost REAL,
            promo_period_revenue REAL, revenue_per_promo_dollar REAL,
            lift_measurable INTEGER
        )
    """)
    conn.execute(
        "INSERT INTO results_trade_efficiency VALUES (?,?,?,?,?,?,?,?)",
        ("Walmart", 0.12, 120000.0, 1000000.0, 50000.0, 400000.0, 8.0, 1),
    )
    conn.execute("""
        CREATE TABLE results_promo_roi (
            promo_id TEXT, sku_id TEXT, retailer_id TEXT, retailer TEXT,
            start_week TEXT, end_week TEXT, promo_cost REAL, promo_type TEXT,
            has_sufficient_baseline INTEGER, baseline_weekly_revenue REAL,
            promo_revenue REAL, promo_weeks INTEGER, incremental_revenue REAL,
            incremental_margin REAL, is_money_losing INTEGER
        )
    """)
    conn.execute(
        "INSERT INTO results_promo_roi VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        ("P001", "SKU-001", "walmart", "Walmart", "2024-01-01", "2024-01-14",
         5000.0, "TPR", 1, 20000.0, 30000.0, 2, 10000.0, 7000.0, 0),
    )
    conn.execute("""
        CREATE TABLE results_accrual (
            month TEXT, accrued REAL, actual REAL, variance REAL
        )
    """)
    conn.execute(
        "INSERT INTO results_accrual VALUES (?,?,?,?)",
        ("2024-01", 100000.0, 95000.0, 5000.0),
    )
    conn.commit()
    conn.close()


@pytest.fixture
def full_results_db(tmp_path):
    db_path = tmp_path / "results.db"
    _make_results_db(db_path)
    return db_path


@pytest.fixture
def partial_results_db(tmp_path):
    """Only Move 1 and Move 3 tables populated."""
    db_path = tmp_path / "results.db"
    conn = sqlite3.connect(db_path)
    conn.execute("""
        CREATE TABLE results_net_revenue (
            retailer TEXT, gross_revenue REAL, trade_spend REAL,
            net_revenue REAL, net_to_gross_ratio REAL
        )
    """)
    conn.execute(
        "INSERT INTO results_net_revenue VALUES (?,?,?,?,?)",
        ("Walmart", 1000000.0, 120000.0, 880000.0, 0.88),
    )
    conn.execute("""
        CREATE TABLE results_leakage_summary (
            leakage_type TEXT, display_name TEXT, dollar_total REAL,
            instance_count INTEGER, classification TEXT
        )
    """)
    conn.execute(
        "INSERT INTO results_leakage_summary VALUES (?,?,?,?,?)",
        ("double_funded", "Double-Funded", 15264.0, 173, "Recoverable"),
    )
    conn.execute("""
        CREATE TABLE results_leakage_instances (
            leakage_type TEXT, deduction_id TEXT, retailer_id TEXT,
            promo_id TEXT, period TEXT, agreed_amount REAL,
            actual_amount REAL, variance REAL, classification TEXT
        )
    """)
    conn.commit()
    conn.close()
    return db_path


# ---------------------------------------------------------------------------
# AE3 — sheet names and order
# ---------------------------------------------------------------------------

def test_workbook_has_correct_sheet_names_and_order(full_results_db):
    """AE3: generate_workbook returns xlsx with the six sheets in canonical order."""
    wb_bytes = generate_workbook(full_results_db)
    assert isinstance(wb_bytes, bytes)
    assert len(wb_bytes) > 0

    wb = load_workbook(BytesIO(wb_bytes))
    assert wb.sheetnames == _EXPECTED_SHEETS


def test_workbook_returns_bytes_from_missing_db(tmp_path):
    """generate_workbook handles missing results.db without raising."""
    wb_bytes = generate_workbook(tmp_path / "nonexistent.db")
    wb = load_workbook(BytesIO(wb_bytes))
    assert wb.sheetnames == _EXPECTED_SHEETS


# ---------------------------------------------------------------------------
# Each sheet has data
# ---------------------------------------------------------------------------

def test_each_sheet_has_at_least_one_data_row(full_results_db):
    """Happy path: each sheet is not empty (has content beyond row 1)."""
    wb = load_workbook(BytesIO(generate_workbook(full_results_db)))
    for sheet_name in _EXPECTED_SHEETS:
        ws = wb[sheet_name]
        # max_row includes headers and content
        assert ws.max_row > 1, f"Sheet '{sheet_name}' appears empty (max_row={ws.max_row})"


# ---------------------------------------------------------------------------
# Numbers match results.db (R9)
# ---------------------------------------------------------------------------

def test_net_revenue_sheet_totals_match_results_db(full_results_db):
    """Happy path: Net Revenue Ranking sheet totals equal what's in results.db."""
    wb = load_workbook(BytesIO(generate_workbook(full_results_db)))
    ws = wb["Net Revenue Ranking"]

    # Row 7 is the header row (rows 1-5 = header, row 6 = section title, row 7 = table headers)
    # Data row is row 8 for the single fixture row; totals row is row 9.
    # We look for the "Total" label and read its gross revenue column value.
    total_row = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "Total":
                total_row = cell.row
                break
        if total_row:
            break

    assert total_row is not None, "No 'Total' row found in Net Revenue Ranking sheet"

    # Column C (index 3) is Gross Revenue total
    gross_total_in_sheet = ws.cell(row=total_row, column=3).value
    assert gross_total_in_sheet == pytest.approx(1000000.0, rel=1e-3)


# ---------------------------------------------------------------------------
# Partial build state — placeholder, not error
# ---------------------------------------------------------------------------

def test_partial_build_state_produces_valid_workbook_with_placeholders(partial_results_db):
    """Edge case: workbook generated when only Move 1 and Move 3 are populated."""
    wb_bytes = generate_workbook(partial_results_db)
    wb = load_workbook(BytesIO(wb_bytes))

    # All six sheets still present
    assert wb.sheetnames == _EXPECTED_SHEETS

    # Net Revenue Ranking has real data
    ws_nr = wb["Net Revenue Ranking"]
    text_values = [c.value for row in ws_nr.iter_rows() for c in row if c.value]
    assert "Walmart" in text_values

    # Trade Spend Efficiency and other missing sheets contain the placeholder text
    for sheet_name in ["Trade Spend Efficiency", "Promotional ROI", "Accrual Reconciliation"]:
        ws = wb[sheet_name]
        all_values = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
        has_placeholder = any("Not yet computed" in v for v in all_values)
        assert has_placeholder, f"'{sheet_name}' missing placeholder text for partial build state"
