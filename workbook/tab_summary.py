"""Tab 1: Summary — key metrics across all five analytical moves."""

from __future__ import annotations

import sqlite3
from workbook.styles import DEFAULT_BUILT_DATE
from pathlib import Path

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from workbook.styles import (
    ALIGN_CENTER,
    ALIGN_LEFT,
    ALIGN_RIGHT,
    BORDER_SECTION,
    FONT_BODY,
    FONT_HEADER,
    FONT_KPI_LABEL,
    FONT_KPI_VALUE,
    FONT_SECTION,
    FONT_SMALL,
    NUM_FMT_DOLLAR,
    NUM_FMT_PCT,
    TAB_NAMES,
    FONT_NAV,
)
from workbook.windows import read_trailing_months


def _read(db_path: Path) -> dict:
    if not db_path.exists():
        return {}
    conn = sqlite3.connect(db_path)
    try:
        data: dict = {}

        try:
            rows = conn.execute(
                "SELECT retailer, gross_revenue, trade_spend, net_revenue, net_to_gross_ratio "
                "FROM results_net_revenue ORDER BY net_revenue DESC"
            ).fetchall()
            data["net_revenue"] = rows
        except Exception:
            data["net_revenue"] = None

        try:
            rows = conn.execute(
                "SELECT leakage_type, display_name, dollar_total, instance_count, classification "
                "FROM results_leakage_summary"
            ).fetchall()
            data["leakage_summary"] = rows
        except Exception:
            data["leakage_summary"] = None

        try:
            rows = conn.execute(
                "SELECT retailer, trade_spend_pct FROM results_trade_efficiency"
            ).fetchall()
            data["efficiency"] = rows
        except Exception:
            data["efficiency"] = None

        try:
            row = conn.execute(
                "SELECT COUNT(*), SUM(CASE WHEN is_money_losing = 1 THEN 1 ELSE 0 END) "
                "FROM results_promo_roi WHERE has_sufficient_baseline = 1"
            ).fetchone()
            data["promo_roi"] = row
        except Exception:
            data["promo_roi"] = None

        try:
            row = conn.execute(
                "SELECT SUM(accrued), SUM(actual), SUM(variance) FROM results_accrual"
            ).fetchone()
            data["accrual"] = row
        except Exception:
            data["accrual"] = None

        return data
    finally:
        conn.close()


def _placeholder(ws: Worksheet, row: int, col: int = 2) -> None:
    cell = ws.cell(row=row, column=col, value="Not yet computed — run the pipeline first.")
    cell.font = FONT_SMALL


def build_summary(ws: Worksheet, db_path: Path, built_date=None) -> None:
    data = _read(db_path)
    # Month span derived from the pipeline output, never hardcoded.
    _months = read_trailing_months(db_path)

    ws.sheet_view.showGridLines = False

    col_widths = [3, 28, 18, 18, 18, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- Header ---
    ws.merge_cells("B1:F1")
    ws["B1"] = "Cinderhaven Provisions"
    ws["B1"].font = FONT_HEADER

    ws.merge_cells("B2:F2")
    ws["B2"] = "Trade Spend Leakage — Executive Summary"
    ws["B2"].font = FONT_SECTION

    ws.merge_cells("B3:F3")
    ws["B3"] = f"Built {(built_date or DEFAULT_BUILT_DATE).isoformat()}"
    ws["B3"].font = FONT_SMALL

    for col in range(2, 7):
        ws.cell(row=4, column=col).border = BORDER_SECTION

    row = 6

    # --- Move 1: Net Revenue Ranking KPIs ---
    ws.cell(row=row, column=2, value="01  Net Revenue Ranking").font = FONT_SECTION
    row += 1

    if data.get("net_revenue"):
        nr = data["net_revenue"]
        total_gross = sum(r[1] for r in nr)
        total_trade = sum(r[2] for r in nr)
        total_net = sum(r[3] for r in nr)
        blended_rate = total_trade / total_gross if total_gross else 0.0

        kpis = [
            ("Gross Revenue", total_gross, NUM_FMT_DOLLAR),
            ("Total Trade Spend", total_trade, NUM_FMT_DOLLAR),
            ("Net Revenue", total_net, NUM_FMT_DOLLAR),
            ("Blended Trade Rate", blended_rate, NUM_FMT_PCT),
        ]
        for c_off, (label, value, fmt) in enumerate(kpis):
            val_cell = ws.cell(row=row, column=2 + c_off, value=value)
            val_cell.font = FONT_KPI_VALUE
            val_cell.number_format = fmt
            val_cell.alignment = ALIGN_CENTER
            lbl_cell = ws.cell(row=row + 1, column=2 + c_off, value=label)
            lbl_cell.font = FONT_KPI_LABEL
            lbl_cell.alignment = ALIGN_CENTER
        row += 2
    else:
        _placeholder(ws, row)
        row += 1

    for col in range(2, 7):
        ws.cell(row=row, column=col).border = BORDER_SECTION
    row += 2

    # --- Move 3: Leakage Detection KPIs ---
    ws.cell(row=row, column=2, value="03  Leakage Detection").font = FONT_SECTION
    row += 1

    if data.get("leakage_summary"):
        ls = data["leakage_summary"]
        total_leakage = sum(r[2] for r in ls)
        total_instances = sum(r[3] for r in ls)

        val_cell = ws.cell(row=row, column=2, value=total_leakage)
        val_cell.font = FONT_KPI_VALUE
        val_cell.number_format = NUM_FMT_DOLLAR
        val_cell.alignment = ALIGN_CENTER
        ws.cell(row=row + 1, column=2, value="Total Leakage").font = FONT_KPI_LABEL
        ws.cell(row=row + 1, column=2).alignment = ALIGN_CENTER

        val2 = ws.cell(row=row, column=3, value=total_instances)
        val2.font = FONT_KPI_VALUE
        val2.alignment = ALIGN_CENTER
        ws.cell(row=row + 1, column=3, value="Instances").font = FONT_KPI_LABEL
        ws.cell(row=row + 1, column=3).alignment = ALIGN_CENTER

        row += 2

        # Sub-type breakdown
        sub_headers = ["Type", "Instances", "Amount", "Classification"]
        for c, h in enumerate(sub_headers, 2):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = FONT_BODY
            cell.alignment = ALIGN_CENTER
        row += 1
        for ltype, display_name, dollar_total, instance_count, classification in ls:
            ws.cell(row=row, column=2, value=display_name).font = FONT_BODY
            c_inst = ws.cell(row=row, column=3, value=instance_count)
            c_inst.font = FONT_BODY
            c_inst.alignment = ALIGN_RIGHT
            c_amt = ws.cell(row=row, column=4, value=dollar_total)
            c_amt.font = FONT_BODY
            c_amt.number_format = NUM_FMT_DOLLAR
            c_amt.alignment = ALIGN_RIGHT
            ws.cell(row=row, column=5, value=classification).font = FONT_BODY
            row += 1
    else:
        _placeholder(ws, row)
        row += 1

    for col in range(2, 7):
        ws.cell(row=row, column=col).border = BORDER_SECTION
    row += 2

    # --- Move 2: Efficiency ---
    ws.cell(row=row, column=2, value="02  Trade Spend Efficiency").font = FONT_SECTION
    row += 1

    if data.get("efficiency"):
        eff = data["efficiency"]
        avg_rate = sum(r[1] for r in eff) / len(eff) if eff else 0.0
        val_cell = ws.cell(row=row, column=2, value=avg_rate)
        val_cell.font = FONT_KPI_VALUE
        val_cell.number_format = NUM_FMT_PCT
        val_cell.alignment = ALIGN_CENTER
        ws.cell(row=row + 1, column=2, value="Avg Trade Spend Rate").font = FONT_KPI_LABEL
        ws.cell(row=row + 1, column=2).alignment = ALIGN_CENTER
        row += 2
    else:
        _placeholder(ws, row)
        row += 1

    for col in range(2, 7):
        ws.cell(row=row, column=col).border = BORDER_SECTION
    row += 2

    # --- Move 4: Promo ROI ---
    ws.cell(row=row, column=2, value="04  Promotional ROI").font = FONT_SECTION
    row += 1

    if data.get("promo_roi") and data["promo_roi"][0] is not None:
        measurable, money_losing = data["promo_roi"]
        money_losing = money_losing or 0

        val_cell = ws.cell(row=row, column=2, value=measurable)
        val_cell.font = FONT_KPI_VALUE
        val_cell.alignment = ALIGN_CENTER
        ws.cell(row=row + 1, column=2, value="Measurable Promos").font = FONT_KPI_LABEL
        ws.cell(row=row + 1, column=2).alignment = ALIGN_CENTER

        val2 = ws.cell(row=row, column=3, value=money_losing)
        val2.font = FONT_KPI_VALUE
        val2.alignment = ALIGN_CENTER
        ws.cell(row=row + 1, column=3, value="Money-Losing").font = FONT_KPI_LABEL
        ws.cell(row=row + 1, column=3).alignment = ALIGN_CENTER

        row += 2
    else:
        _placeholder(ws, row)
        row += 1

    for col in range(2, 7):
        ws.cell(row=row, column=col).border = BORDER_SECTION
    row += 2

    # --- Move 5: Accrual ---
    ws.cell(row=row, column=2, value="05  Accrual Reconciliation").font = FONT_SECTION
    row += 1

    if data.get("accrual") and data["accrual"][0] is not None:
        total_accrued, total_actual, total_variance = data["accrual"]
        kpis = [
            ("Total Accrued", total_accrued, NUM_FMT_DOLLAR),
            ("Total Actual", total_actual, NUM_FMT_DOLLAR),
            ("Net Variance", total_variance, NUM_FMT_DOLLAR),
        ]
        for c_off, (label, value, fmt) in enumerate(kpis):
            val_cell = ws.cell(row=row, column=2 + c_off, value=value)
            val_cell.font = FONT_KPI_VALUE
            val_cell.number_format = fmt
            val_cell.alignment = ALIGN_CENTER
            lbl_cell = ws.cell(row=row + 1, column=2 + c_off, value=label)
            lbl_cell.font = FONT_KPI_LABEL
            lbl_cell.alignment = ALIGN_CENTER
        row += 2
    else:
        _placeholder(ws, row)
        row += 1

    for col in range(2, 7):
        ws.cell(row=row, column=col).border = BORDER_SECTION
    row += 2

    # --- Navigation ---
    ws.cell(row=row, column=2, value="Navigate to:").font = FONT_SECTION
    row += 1
    for i, tab_name in enumerate(TAB_NAMES):
        cell = ws.cell(row=row, column=2 + i, value=tab_name)
        cell.font = FONT_NAV
        cell.hyperlink = f"#'{tab_name}'!A1"

    row += 2
    ws.merge_cells(f"B{row}:F{row}")
    callout = ws.cell(
        row=row, column=2,
        value=(
            "This workbook summarises "
            + (f"trailing-{_months}-month " if _months else "trailing ")
            + "trade spend for Cinderhaven Provisions. "
            "Numbers match the interactive dashboard. All data sourced from results.db — "
            "the pre-computed output of the trade spend pipeline."
        ),
    )
    callout.font = FONT_SMALL
    callout.alignment = ALIGN_LEFT
