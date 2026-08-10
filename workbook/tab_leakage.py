"""Tab 3: Leakage Detection — four sub-types with full instance-level detail."""

from __future__ import annotations

import sqlite3
from workbook.styles import DEFAULT_BUILT_DATE
from pathlib import Path

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet

from workbook.styles import (
    ALIGN_CENTER,
    ALIGN_LEFT,
    ALIGN_RIGHT,
    BORDER_SECTION,
    FILL_BAD,
    FILL_GOOD,
    FILL_WARN,
    FONT_BODY,
    FONT_HEADER,
    FONT_SECTION,
    FONT_SMALL,
    NUM_FMT_DOLLAR,
    SANS,
    TABLE_STYLE,
)

_PLACEHOLDER = "Not yet computed — run the pipeline first."


def _read(db_path: Path) -> dict | None:
    if not db_path.exists():
        return None
    conn = sqlite3.connect(db_path)
    try:
        summary = conn.execute(
            "SELECT leakage_type, display_name, dollar_total, instance_count, classification "
            "FROM results_leakage_summary"
        ).fetchall()

        instances = conn.execute(
            "SELECT leakage_type, deduction_id, retailer_id, promo_id, "
            "period, agreed_amount, actual_amount, variance, classification "
            "FROM results_leakage_instances ORDER BY leakage_type, actual_amount DESC"
        ).fetchall()

        return {"summary": summary, "instances": instances}
    except Exception:
        return None
    finally:
        conn.close()


def build_leakage(ws: Worksheet, db_path: Path, built_date=None) -> None:
    data = _read(db_path)

    ws.sheet_view.showGridLines = False

    col_widths = [3, 22, 14, 14, 16, 14, 14, 14, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("B1:I1")
    ws["B1"] = "Leakage Detection"
    ws["B1"].font = FONT_HEADER

    ws.merge_cells("B2:I2")
    ws["B2"] = (
        "Four leakage sub-types: double-funded promotions, ghost promos (no matching calendar entry), "
        "rate discrepancies (actual > agreed), and unauthorized deductions. "
        "Deduction IDs link back to the Cinderhaven SSOT."
    )
    ws["B2"].font = FONT_SMALL

    ws.merge_cells("B3:I3")
    ws["B3"] = f"Built {(built_date or DEFAULT_BUILT_DATE).isoformat()}"
    ws["B3"].font = FONT_SMALL

    if data is None:
        ws.cell(row=5, column=2, value=_PLACEHOLDER).font = FONT_SMALL
        return

    # --- Summary ledger ---
    row = 5
    ws.cell(row=row, column=2, value="Leakage Summary").font = FONT_SECTION
    row += 1

    sum_headers = ["Type", "Instances", "Amount", "Classification"]
    for c, h in enumerate(sum_headers, 2):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(name=SANS, size=11, bold=True)
        cell.alignment = ALIGN_CENTER

    sum_start = row + 1
    for ltype, display_name, dollar_total, instance_count, classification in data["summary"]:
        row += 1
        ws.cell(row=row, column=2, value=display_name).font = FONT_BODY
        c_inst = ws.cell(row=row, column=3, value=instance_count)
        c_inst.font = FONT_BODY
        c_inst.alignment = ALIGN_RIGHT
        c_amt = ws.cell(row=row, column=4, value=dollar_total)
        c_amt.number_format = NUM_FMT_DOLLAR
        c_amt.alignment = ALIGN_RIGHT
        ws.cell(row=row, column=5, value=classification).font = FONT_BODY

    sum_end = row

    if data["summary"]:
        sum_table = Table(displayName="tbl_LeakageSummary", ref=f"B{sum_start - 1}:E{sum_end}")
        sum_table.tableStyleInfo = TABLE_STYLE
        ws.add_table(sum_table)

    # Totals row
    row += 1
    ws.cell(row=row, column=2, value="Total").font = Font(name=SANS, size=11, bold=True)
    c_ti = ws.cell(row=row, column=3, value=sum(r[3] for r in data["summary"]))
    c_ti.font = Font(name=SANS, size=11, bold=True)
    c_ti.alignment = ALIGN_RIGHT
    c_ta = ws.cell(row=row, column=4, value=sum(r[2] for r in data["summary"]))
    c_ta.number_format = NUM_FMT_DOLLAR
    c_ta.alignment = ALIGN_RIGHT
    c_ta.font = Font(name=SANS, size=11, bold=True)

    row += 1
    for col in range(2, 10):
        ws.cell(row=row, column=col).border = BORDER_SECTION

    row += 2

    # --- Instance table ---
    ws.cell(row=row, column=2, value="Instance Detail").font = FONT_SECTION
    row += 1

    inst_headers = [
        "Leakage Type", "Deduction ID", "Retailer", "Promo ID",
        "Period", "Agreed Amount", "Actual Amount", "Variance", "Classification",
    ]
    inst_header_row = row
    for c, h in enumerate(inst_headers, 2):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(name=SANS, size=10, bold=True)
        cell.alignment = ALIGN_CENTER

    ws.freeze_panes = f"B{row + 1}"

    for inst_row_data in data["instances"]:
        row += 1
        ltype, ded_id, retailer_id, promo_id, period, agreed, actual, variance, classification = inst_row_data
        ws.cell(row=row, column=2, value=ltype).font = FONT_BODY
        ws.cell(row=row, column=3, value=ded_id).font = FONT_BODY
        ws.cell(row=row, column=4, value=retailer_id).font = FONT_BODY
        ws.cell(row=row, column=5, value=promo_id).font = FONT_BODY
        ws.cell(row=row, column=6, value=period).font = FONT_BODY

        c_agr = ws.cell(row=row, column=7, value=agreed)
        c_agr.number_format = NUM_FMT_DOLLAR
        c_agr.alignment = ALIGN_RIGHT

        c_act = ws.cell(row=row, column=8, value=actual)
        c_act.number_format = NUM_FMT_DOLLAR
        c_act.alignment = ALIGN_RIGHT

        c_var = ws.cell(row=row, column=9, value=variance)
        c_var.number_format = NUM_FMT_DOLLAR
        c_var.alignment = ALIGN_RIGHT

        ws.cell(row=row, column=10, value=classification).font = FONT_BODY

    inst_end = row

    if data["instances"]:
        inst_table = Table(displayName="tbl_LeakageInstances", ref=f"B{inst_header_row}:J{inst_end}")
        inst_table.tableStyleInfo = TABLE_STYLE
        ws.add_table(inst_table)

        class_range = f"J{inst_header_row + 1}:J{inst_end}"
        ws.conditional_formatting.add(
            class_range,
            CellIsRule(operator="equal", formula=['"Recoverable"'], fill=FILL_BAD),
        )
        ws.conditional_formatting.add(
            class_range,
            CellIsRule(operator="equal", formula=['"Reallocatable"'], fill=FILL_WARN),
        )

    row += 2
    ws.merge_cells(f"B{row}:I{row}")
    note = ws.cell(
        row=row, column=2,
        value=(
            "Double-funded: promo_billback matched to off-invoice promotion. "
            "Ghost: promo_billback with no matching promotion ±14 days. "
            "Rate discrepancy: billback > agreed promo_cost by >5%. "
            "Unauthorized: deduction type outside known operational set."
        ),
    )
    note.font = FONT_SMALL
    note.alignment = ALIGN_LEFT
