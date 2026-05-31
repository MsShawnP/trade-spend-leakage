"""Tab 6: Accrual Reconciliation — monthly accrued vs actual trade spend."""

from __future__ import annotations

import sqlite3
from datetime import date
from pathlib import Path

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet

from workbook.styles import (
    ALIGN_CENTER,
    ALIGN_LEFT,
    ALIGN_RIGHT,
    FONT_BODY,
    FONT_HEADER,
    FONT_KPI_LABEL,
    FONT_KPI_VALUE,
    FONT_SECTION,
    FONT_SMALL,
    FILL_BAD,
    FILL_GOOD,
    NUM_FMT_DOLLAR,
    SANS,
    TABLE_STYLE,
)

_PLACEHOLDER = "Not yet computed — run the pipeline first."


def _read(db_path: Path) -> list[tuple] | None:
    if not db_path.exists():
        return None
    conn = sqlite3.connect(db_path)
    try:
        return conn.execute(
            "SELECT month, accrued, actual, variance FROM results_accrual ORDER BY month ASC"
        ).fetchall()
    except Exception:
        return None
    finally:
        conn.close()


def build_accrual(ws: Worksheet, db_path: Path) -> None:
    rows = _read(db_path)

    ws.sheet_view.showGridLines = False

    col_widths = [3, 16, 16, 16, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("B1:E1")
    ws["B1"] = "Accrual Reconciliation"
    ws["B1"].font = FONT_HEADER

    ws.merge_cells("B2:E2")
    ws["B2"] = (
        "Monthly accrued trade spend (rate card × scan revenue) versus actual deductions. "
        "Positive variance = accrued more than was taken (under-billed). "
        "Negative = over-billed against accrual."
    )
    ws["B2"].font = FONT_SMALL

    ws.merge_cells("B3:E3")
    ws["B3"] = f"Built {date.today().isoformat()}"
    ws["B3"].font = FONT_SMALL

    if rows is None:
        ws.cell(row=5, column=2, value=_PLACEHOLDER).font = FONT_SMALL
        return

    # --- KPI summary ---
    row = 5
    total_accrued = sum(r[1] for r in rows)
    total_actual = sum(r[2] for r in rows)
    total_variance = sum(r[3] for r in rows)

    kpis = [
        ("Total Accrued", total_accrued),
        ("Total Actual", total_actual),
        ("Net Variance", total_variance),
    ]
    for c_off, (label, value) in enumerate(kpis):
        val_cell = ws.cell(row=row, column=2 + c_off, value=value)
        val_cell.font = FONT_KPI_VALUE
        val_cell.number_format = NUM_FMT_DOLLAR
        val_cell.alignment = ALIGN_CENTER
        lbl_cell = ws.cell(row=row + 1, column=2 + c_off, value=label)
        lbl_cell.font = FONT_KPI_LABEL
        lbl_cell.alignment = ALIGN_CENTER

    row += 3

    # --- Monthly table ---
    ws.cell(row=row, column=2, value="Monthly Detail (Trailing 12 Months)").font = FONT_SECTION
    row += 1

    headers = ["Month", "Accrued", "Actual Deducted", "Variance"]
    header_row = row
    for c, h in enumerate(headers, 2):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = Font(name=SANS, size=11, bold=True)
        cell.alignment = ALIGN_CENTER

    for i, (month, accrued, actual, variance) in enumerate(rows):
        r = header_row + 1 + i
        ws.cell(row=r, column=2, value=month).font = FONT_BODY

        c_a = ws.cell(row=r, column=3, value=accrued)
        c_a.number_format = NUM_FMT_DOLLAR
        c_a.alignment = ALIGN_RIGHT

        c_ac = ws.cell(row=r, column=4, value=actual)
        c_ac.number_format = NUM_FMT_DOLLAR
        c_ac.alignment = ALIGN_RIGHT

        c_v = ws.cell(row=r, column=5, value=variance)
        c_v.number_format = NUM_FMT_DOLLAR
        c_v.alignment = ALIGN_RIGHT

    table_end = header_row + len(rows)

    tbl = Table(displayName="tbl_Accrual", ref=f"B{header_row}:E{table_end}")
    tbl.tableStyleInfo = TABLE_STYLE
    ws.add_table(tbl)

    # Positive variance = under-billed (good), negative = over-billed (bad)
    var_range = f"E{header_row + 1}:E{table_end}"
    ws.conditional_formatting.add(
        var_range,
        CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=FILL_GOOD),
    )
    ws.conditional_formatting.add(
        var_range,
        CellIsRule(operator="lessThan", formula=["0"], fill=FILL_BAD),
    )

    # Totals row
    totals_row = table_end + 1
    ws.cell(row=totals_row, column=2, value="Total").font = Font(name=SANS, size=11, bold=True)
    c_ta = ws.cell(row=totals_row, column=3, value=total_accrued)
    c_ta.number_format = NUM_FMT_DOLLAR
    c_ta.alignment = ALIGN_RIGHT
    c_ta.font = Font(name=SANS, size=11, bold=True)
    c_tac = ws.cell(row=totals_row, column=4, value=total_actual)
    c_tac.number_format = NUM_FMT_DOLLAR
    c_tac.alignment = ALIGN_RIGHT
    c_tac.font = Font(name=SANS, size=11, bold=True)
    c_tv = ws.cell(row=totals_row, column=5, value=total_variance)
    c_tv.number_format = NUM_FMT_DOLLAR
    c_tv.alignment = ALIGN_RIGHT
    c_tv.font = Font(name=SANS, size=11, bold=True)

    note_row = totals_row + 2
    ws.merge_cells(f"B{note_row}:E{note_row}")
    note = ws.cell(
        row=note_row, column=2,
        value=(
            "Accrued: trailing-12-month scan revenue × structural rate card (sku_costs) per channel. "
            "Actual: all deductions recorded in retailer_deductions, grouped by month."
        ),
    )
    note.font = FONT_SMALL
    note.alignment = ALIGN_LEFT
