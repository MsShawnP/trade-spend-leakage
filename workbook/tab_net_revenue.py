"""Tab 2: Net Revenue Ranking — gross revenue, trade spend, net revenue per retailer."""

from __future__ import annotations

import sqlite3
from datetime import date
from pathlib import Path

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet

from workbook.styles import (
    ALIGN_CENTER,
    ALIGN_LEFT,
    ALIGN_RIGHT,
    FILL_BAD,
    FILL_GOOD,
    FONT_BODY,
    FONT_HEADER,
    FONT_SECTION,
    FONT_SMALL,
    NUM_FMT_DOLLAR,
    NUM_FMT_PCT,
    SANS,
    TABLE_STYLE,
)

_PLACEHOLDER = "Not yet computed — run the pipeline first."


def _read(db_path: Path) -> list[tuple] | None:
    if not db_path.exists():
        return None
    conn = sqlite3.connect(db_path)
    try:
        return conn.execute(
            "SELECT retailer, gross_revenue, trade_spend, net_revenue, net_to_gross_ratio "
            "FROM results_net_revenue ORDER BY net_revenue DESC"
        ).fetchall()
    except Exception:
        return None
    finally:
        conn.close()


def build_net_revenue(ws: Worksheet, db_path: Path) -> None:
    rows = _read(db_path)

    ws.sheet_view.showGridLines = False

    col_widths = [3, 24, 18, 18, 18, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("B1:F1")
    ws["B1"] = "Net Revenue Ranking"
    ws["B1"].font = FONT_HEADER

    ws.merge_cells("B2:F2")
    ws["B2"] = (
        "Retailers ranked by net revenue after structural trade spend. "
        "Gross and net rank may differ — the crossing lines in the dashboard tell the story."
    )
    ws["B2"].font = FONT_SMALL

    ws.merge_cells("B3:F3")
    ws["B3"] = f"Built {date.today().isoformat()}"
    ws["B3"].font = FONT_SMALL

    if rows is None:
        ws.cell(row=5, column=2, value=_PLACEHOLDER).font = FONT_SMALL
        return

    ws.cell(row=5, column=2, value="Trailing 52 Weeks").font = FONT_SECTION

    headers = ["Retailer", "Gross Revenue", "Trade Spend", "Net Revenue", "Net-to-Gross %"]
    header_row = 6
    for c, h in enumerate(headers, 2):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = Font(name=SANS, size=11, bold=True)
        cell.alignment = ALIGN_CENTER

    for i, (retailer, gross, trade, net, ratio) in enumerate(rows):
        r = header_row + 1 + i
        ws.cell(row=r, column=2, value=retailer).font = FONT_BODY

        c_gross = ws.cell(row=r, column=3, value=gross)
        c_gross.number_format = NUM_FMT_DOLLAR
        c_gross.alignment = ALIGN_RIGHT

        c_trade = ws.cell(row=r, column=4, value=trade)
        c_trade.number_format = NUM_FMT_DOLLAR
        c_trade.alignment = ALIGN_RIGHT

        c_net = ws.cell(row=r, column=5, value=net)
        c_net.number_format = NUM_FMT_DOLLAR
        c_net.alignment = ALIGN_RIGHT

        c_ratio = ws.cell(row=r, column=6, value=ratio)
        c_ratio.number_format = NUM_FMT_PCT
        c_ratio.alignment = ALIGN_CENTER

    table_end = header_row + len(rows)

    tbl = Table(displayName="tbl_NetRevenue", ref=f"B{header_row}:F{table_end}")
    tbl.tableStyleInfo = TABLE_STYLE
    ws.add_table(tbl)

    # Conditional formatting on net-to-gross ratio: >= 0.83 is good (trade rate <= 17%)
    ratio_range = f"F{header_row + 1}:F{table_end}"
    ws.conditional_formatting.add(
        ratio_range,
        CellIsRule(operator="greaterThanOrEqual", formula=["0.83"], fill=FILL_GOOD),
    )
    ws.conditional_formatting.add(
        ratio_range,
        CellIsRule(operator="lessThan", formula=["0.83"], fill=FILL_BAD),
    )

    # Totals row
    totals_row = table_end + 1
    ws.cell(row=totals_row, column=2, value="Total").font = Font(name=SANS, size=11, bold=True)
    c_tg = ws.cell(row=totals_row, column=3, value=sum(r[1] for r in rows))
    c_tg.number_format = NUM_FMT_DOLLAR
    c_tg.alignment = ALIGN_RIGHT
    c_tg.font = Font(name=SANS, size=11, bold=True)
    c_tt = ws.cell(row=totals_row, column=4, value=sum(r[2] for r in rows))
    c_tt.number_format = NUM_FMT_DOLLAR
    c_tt.alignment = ALIGN_RIGHT
    c_tt.font = Font(name=SANS, size=11, bold=True)
    c_tn = ws.cell(row=totals_row, column=5, value=sum(r[3] for r in rows))
    c_tn.number_format = NUM_FMT_DOLLAR
    c_tn.alignment = ALIGN_RIGHT
    c_tn.font = Font(name=SANS, size=11, bold=True)

    note_row = totals_row + 2
    ws.merge_cells(f"B{note_row}:F{note_row}")
    note = ws.cell(
        row=note_row, column=2,
        value=(
            "Structural trade spend rate from sku_costs rate card per channel. "
            "Regional chains use the blended regional rate. Trailing 52 weeks of scan data."
        ),
    )
    note.font = FONT_SMALL
    note.alignment = ALIGN_LEFT
