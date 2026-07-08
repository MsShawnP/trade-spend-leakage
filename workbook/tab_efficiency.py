"""Tab 4: Trade Spend Efficiency — total trade spend rate and promo revenue per dollar."""

from __future__ import annotations

import sqlite3
from datetime import date
from pathlib import Path

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet

from workbook.styles import (
    ALIGN_CENTER,
    ALIGN_LEFT,
    ALIGN_RIGHT,
    FONT_BODY,
    FONT_HEADER,
    FONT_SECTION,
    FONT_SMALL,
    NUM_FMT_DOLLAR,
    NUM_FMT_PCT,
    SANS,
    TABLE_STYLE,
)

_PLACEHOLDER = "Not yet computed — run the pipeline first."


def _read(db_path: Path) -> list[tuple] | None:
    if not db_path.exists():
        return None
    conn = sqlite3.connect(db_path)
    try:
        return conn.execute(
            "SELECT retailer, trade_spend_pct, trade_spend, gross_revenue, "
            "total_promo_cost, promo_period_revenue, revenue_per_promo_dollar, lift_measurable "
            "FROM results_trade_efficiency ORDER BY trade_spend_pct ASC"
        ).fetchall()
    except Exception:
        return None
    finally:
        conn.close()


def build_efficiency(ws: Worksheet, db_path: Path) -> None:
    rows = _read(db_path)

    ws.sheet_view.showGridLines = False

    col_widths = [3, 22, 14, 14, 14, 14, 16, 16, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("B1:I1")
    ws["B1"] = "Trade Spend Efficiency"
    ws["B1"].font = FONT_HEADER

    ws.merge_cells("B2:I2")
    ws["B2"] = (
        "Total trade spend per retailer (share of gross revenue consumed by structural "
        "rate-card discounts plus operational deductions) and revenue generated per "
        "promotional dollar."
    )
    ws["B2"].font = FONT_SMALL

    ws.merge_cells("B3:I3")
    ws["B3"] = f"Built {date.today().isoformat()}"
    ws["B3"].font = FONT_SMALL

    if rows is None:
        ws.cell(row=5, column=2, value=_PLACEHOLDER).font = FONT_SMALL
        return

    row = 5
    ws.cell(row=row, column=2, value="Per-Retailer Trade Efficiency").font = FONT_SECTION
    row += 1

    headers = [
        "Retailer", "Total Trade Spend %", "Trade Spend $", "Gross Revenue",
        "Total Promo Cost", "Promo Period Revenue", "Revenue / Promo $", "Lift Measurable",
    ]
    header_row = row
    for c, h in enumerate(headers, 2):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = Font(name=SANS, size=11, bold=True)
        cell.alignment = ALIGN_CENTER

    for i, (retailer, trade_pct, trade_spend, gross_rev,
            total_promo_cost, promo_period_rev, rev_per_dollar, lift_measurable) in enumerate(rows):
        r = header_row + 1 + i
        ws.cell(row=r, column=2, value=retailer).font = FONT_BODY

        c_pct = ws.cell(row=r, column=3, value=trade_pct)
        c_pct.number_format = NUM_FMT_PCT
        c_pct.alignment = ALIGN_CENTER

        c_ts = ws.cell(row=r, column=4, value=trade_spend)
        c_ts.number_format = NUM_FMT_DOLLAR
        c_ts.alignment = ALIGN_RIGHT

        c_gr = ws.cell(row=r, column=5, value=gross_rev)
        c_gr.number_format = NUM_FMT_DOLLAR
        c_gr.alignment = ALIGN_RIGHT

        c_pc = ws.cell(row=r, column=6, value=total_promo_cost)
        c_pc.number_format = NUM_FMT_DOLLAR
        c_pc.alignment = ALIGN_RIGHT

        c_pr = ws.cell(row=r, column=7, value=promo_period_rev)
        c_pr.number_format = NUM_FMT_DOLLAR
        c_pr.alignment = ALIGN_RIGHT

        c_rpd = ws.cell(row=r, column=8, value=rev_per_dollar)
        c_rpd.number_format = '0.00'
        c_rpd.alignment = ALIGN_CENTER

        ws.cell(row=r, column=9, value="Yes" if lift_measurable else "No").alignment = ALIGN_CENTER

    table_end = header_row + len(rows)

    tbl = Table(displayName="tbl_TradeEfficiency", ref=f"B{header_row}:I{table_end}")
    tbl.tableStyleInfo = TABLE_STYLE
    ws.add_table(tbl)

    note_row = table_end + 2
    ws.merge_cells(f"B{note_row}:I{note_row}")
    note = ws.cell(
        row=note_row, column=2,
        value=(
            "Total trade spend % = structural rate-card spend plus operational deductions "
            "(damaged, spoilage, late delivery, fines), trailing 52 weeks, from Move 1. "
            "Revenue per promo dollar: total scan revenue during promotional periods ÷ total promo cost."
        ),
    )
    note.font = FONT_SMALL
    note.alignment = ALIGN_LEFT
