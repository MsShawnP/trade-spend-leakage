"""Tab 5: Promotional ROI — per-promo incremental revenue vs cost."""

from __future__ import annotations

import sqlite3
from datetime import date
from pathlib import Path

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet

from workbook.styles import (
    ALIGN_CENTER,
    ALIGN_LEFT,
    ALIGN_RIGHT,
    FILL_BAD,
    FILL_GOOD,
    FILL_WARN,
    FONT_BODY,
    FONT_HEADER,
    FONT_SECTION,
    FONT_SMALL,
    NUM_FMT_DOLLAR,
    SANS,
    TABLE_STYLE,
)

_PLACEHOLDER = "Not yet computed — run the pipeline first."


def _read(db_path: Path) -> list[tuple] | None:
    if not db_path.exists():
        return None
    conn = sqlite3.connect(db_path)
    try:
        return conn.execute(
            "SELECT promo_id, sku_id, retailer, start_week, end_week, promo_cost, promo_type, "
            "has_sufficient_baseline, baseline_weekly_revenue, promo_revenue, promo_weeks, "
            "incremental_revenue, incremental_margin, is_money_losing "
            "FROM results_promo_roi ORDER BY promo_cost DESC"
        ).fetchall()
    except Exception:
        return None
    finally:
        conn.close()


def build_promo_roi(ws: Worksheet, db_path: Path) -> None:
    rows = _read(db_path)

    ws.sheet_view.showGridLines = False

    col_widths = [3, 14, 14, 20, 12, 12, 14, 14, 14, 14, 12, 16, 16, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("B1:N1")
    ws["B1"] = "Promotional ROI"
    ws["B1"].font = FONT_HEADER

    ws.merge_cells("B2:N2")
    ws["B2"] = (
        "Per-promotion incremental revenue versus cost. Baseline from 8-week rolling median of "
        "weekly scan revenue for the promoted SKU × retailer. "
        "Promotions without sufficient baseline data are flagged — they are not failures."
    )
    ws["B2"].font = FONT_SMALL

    ws.merge_cells("B3:N3")
    ws["B3"] = f"Built {date.today().isoformat()}"
    ws["B3"].font = FONT_SMALL

    if rows is None:
        ws.cell(row=5, column=2, value=_PLACEHOLDER).font = FONT_SMALL
        return

    measurable = sum(1 for r in rows if r[7])
    total = len(rows)
    money_losing = sum(1 for r in rows if r[13] == 1)

    row = 5
    ws.cell(
        row=row, column=2,
        value=f"{measurable} of {total} promotions measurable | {money_losing} money-losing",
    ).font = FONT_SECTION
    row += 1

    headers = [
        "Promo ID", "SKU", "Retailer", "Start", "End", "Cost",
        "Type", "Has Baseline", "Baseline Rev/wk", "Promo Revenue",
        "Promo Weeks", "Incremental Rev", "Incremental Margin", "Money-Losing",
    ]
    header_row = row
    for c, h in enumerate(headers, 2):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = Font(name=SANS, size=10, bold=True)
        cell.alignment = ALIGN_CENTER

    ws.freeze_panes = f"B{header_row + 1}"

    def _sort_key(r):
        # Measurable rows first, then money-losing, then by cost desc
        return (0 if r[7] else 1, 0 if r[13] == 1 else 1, -(r[5] or 0))

    sorted_rows = sorted(rows, key=_sort_key)

    for i, row_data in enumerate(sorted_rows):
        (promo_id, sku_id, retailer, start_week, end_week, promo_cost, promo_type,
         has_baseline, baseline_rev, promo_revenue, promo_weeks,
         incr_rev, incr_margin, is_money_losing) = row_data

        r = header_row + 1 + i
        ws.cell(row=r, column=2, value=promo_id).font = FONT_BODY
        ws.cell(row=r, column=3, value=sku_id).font = FONT_BODY
        ws.cell(row=r, column=4, value=retailer).font = FONT_BODY
        ws.cell(row=r, column=5, value=start_week).font = FONT_BODY
        ws.cell(row=r, column=6, value=end_week).font = FONT_BODY

        c_cost = ws.cell(row=r, column=7, value=promo_cost)
        c_cost.number_format = NUM_FMT_DOLLAR
        c_cost.alignment = ALIGN_RIGHT

        ws.cell(row=r, column=8, value=promo_type).font = FONT_BODY
        ws.cell(row=r, column=9, value="Yes" if has_baseline else "No").alignment = ALIGN_CENTER

        c_bl = ws.cell(row=r, column=10, value=baseline_rev)
        c_bl.number_format = NUM_FMT_DOLLAR
        c_bl.alignment = ALIGN_RIGHT

        c_pr = ws.cell(row=r, column=11, value=promo_revenue)
        c_pr.number_format = NUM_FMT_DOLLAR
        c_pr.alignment = ALIGN_RIGHT

        ws.cell(row=r, column=12, value=promo_weeks).alignment = ALIGN_CENTER

        c_ir = ws.cell(row=r, column=13, value=incr_rev)
        c_ir.number_format = NUM_FMT_DOLLAR
        c_ir.alignment = ALIGN_RIGHT

        c_im = ws.cell(row=r, column=14, value=incr_margin)
        c_im.number_format = NUM_FMT_DOLLAR
        c_im.alignment = ALIGN_RIGHT

        if is_money_losing is None:
            ws.cell(row=r, column=15, value="N/A").alignment = ALIGN_CENTER
        else:
            ws.cell(row=r, column=15, value="Yes" if is_money_losing else "No").alignment = ALIGN_CENTER

    table_end = header_row + len(sorted_rows)

    tbl = Table(displayName="tbl_PromoROI", ref=f"B{header_row}:O{table_end}")
    tbl.tableStyleInfo = TABLE_STYLE
    ws.add_table(tbl)

    ml_range = f"O{header_row + 1}:O{table_end}"
    ws.conditional_formatting.add(
        ml_range,
        CellIsRule(operator="equal", formula=['"Yes"'], fill=FILL_BAD),
    )
    ws.conditional_formatting.add(
        ml_range,
        CellIsRule(operator="equal", formula=['"No"'], fill=FILL_GOOD),
    )
    ws.conditional_formatting.add(
        ml_range,
        CellIsRule(operator="equal", formula=['"N/A"'], fill=FILL_WARN),
    )

    note_row = table_end + 2
    ws.merge_cells(f"B{note_row}:N{note_row}")
    note = ws.cell(
        row=note_row, column=2,
        value=(
            "Baseline: 8-week rolling median of weekly scan revenue for promoted SKU × retailer. "
            "Promotions without sufficient pre-promotion history are marked 'No' in Has Baseline — "
            "they are excluded from ROI assessment, not classified as failures."
        ),
    )
    note.font = FONT_SMALL
    note.alignment = ALIGN_LEFT
